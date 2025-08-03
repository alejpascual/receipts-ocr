"""Excel export functionality for transactions and review data."""

import logging
from typing import List, Dict, Any, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

try:
    from .review import ReviewItem
except ImportError:
    from review import ReviewItem

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export transaction data and review items to Excel with beautiful styling."""
    
    def __init__(self, output_path: Path):
        """
        Initialize Excel exporter.
        
        Args:
            output_path: Path for the output Excel file
        """
        self.output_path = output_path
        self.workbook = Workbook()
        
        # 🎨 Professional Color Palette (30 years of Excel design experience!)
        self.colors = {
            # Headers - Deep navy gradient for professional elegance
            'header_bg': '1E3A8A',      # Deep navy blue
            'header_text': 'FFFFFF',     # Pure white
            
            # Status colors - Sophisticated, not garish
            'ok_bg': '10B981',           # Emerald green (success)
            'ok_text': 'FFFFFF',         # White text
            'review_bg': 'DC2626',       # Beautiful red (needs attention)
            'review_text': 'FFFFFF',     # White text
            
            # Section headers
            'section_bg': '6366F1',      # Modern indigo
            'section_text': 'FFFFFF',    # White text
            
            # Summary section
            'summary_bg': 'F3F4F6',      # Light gray
            'summary_accent': '4F46E5',  # Rich purple
            
            # Data rows - Subtle alternating
            'row_even': 'FFFFFF',        # Pure white
            'row_odd': 'F9FAFB',        # Very light gray
            
            # Borders
            'border_color': 'D1D5DB',    # Soft gray
        }
        
        # Professional styling presets
        self.styles = {
            'thin_border': Border(
                left=Side(style='thin', color=self.colors['border_color']),
                right=Side(style='thin', color=self.colors['border_color']),
                top=Side(style='thin', color=self.colors['border_color']),
                bottom=Side(style='thin', color=self.colors['border_color'])
            )
        }
        
    def export_transactions(self, 
                          transactions: List[Dict[str, Any]], 
                          review_items: List[ReviewItem],
                          include_summary: bool = False):
        """
        Export transactions and review data to a single consolidated Excel sheet.
        
        Args:
            transactions: List of transaction dictionaries
            review_items: List of items needing review
            include_summary: Whether to include summary section (always True now)
        """
        try:
            # Remove default sheet
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])
            
            # Create single consolidated sheet
            self._create_consolidated_sheet(transactions, review_items, include_summary)
            
            # Save workbook
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(str(self.output_path))
            
            logger.info(f"Excel file exported to: {self.output_path}")
            
        except Exception as e:
            logger.error(f"Failed to export Excel file: {e}")
            raise
    
    def _create_consolidated_sheet(self, transactions: List[Dict[str, Any]], review_items: List[ReviewItem], include_summary: bool = True):
        """Create a single consolidated sheet with transactions, review items, and summary."""
        ws = self.workbook.create_sheet("All Transactions")
        
        current_row = 1
        
        # Add summary section first if requested
        if include_summary:
            current_row = self._add_summary_section(ws, transactions, current_row)
            current_row += 2  # Add spacing
        
        # Create review lookup for merging data
        review_lookup = {Path(item.file_path).name: item for item in review_items}
        
        # Define headers for consolidated view
        headers = ["File Name", "Date", "Amount", "Category", "Description", 
                  "Review Status", "Review Reason", "Raw Snippet"]
        
        # Add stunning section title
        title_cell = ws.cell(row=current_row, column=1, value="ALL TRANSACTIONS")
        title_cell.font = Font(bold=True, size=16, color=self.colors['section_text'])
        title_cell.fill = PatternFill(start_color=self.colors['section_bg'], end_color=self.colors['section_bg'], fill_type="solid")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        title_cell.border = self.styles['thin_border']
        # Merge title across all columns for dramatic effect
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        current_row += 2
        
        # Add gorgeous professional headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, size=11, color=self.colors['header_text'])
            cell.fill = PatternFill(start_color=self.colors['header_bg'], end_color=self.colors['header_bg'], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.styles['thin_border']
            
            # Auto-adjust column widths for perfect readability
            if header == "File Name":
                ws.column_dimensions[chr(64 + col)].width = 25
            elif header == "Date":
                ws.column_dimensions[chr(64 + col)].width = 12
            elif header == "Amount":
                ws.column_dimensions[chr(64 + col)].width = 12
            elif header == "Category":
                ws.column_dimensions[chr(64 + col)].width = 18
            elif header == "Description":
                ws.column_dimensions[chr(64 + col)].width = 30
            elif header == "Review Status":
                ws.column_dimensions[chr(64 + col)].width = 12
            elif header == "Review Reason":
                ws.column_dimensions[chr(64 + col)].width = 25
            elif header == "Raw Snippet":
                ws.column_dimensions[chr(64 + col)].width = 35
        
        current_row += 1
        
        # Sort transactions: OK first (by filename), then REVIEW (by filename)
        # Separate transactions into OK and REVIEW categories
        ok_transactions = []
        review_transactions = []
        
        for transaction in transactions:
            file_name = transaction.get('file_name', '')
            review_item = review_lookup.get(file_name)
            
            if review_item:
                review_transactions.append((transaction, review_item))
            else:
                ok_transactions.append((transaction, None))
        
        # Sort both lists by filename for consistent cross-checking order
        ok_transactions.sort(key=lambda x: x[0].get('file_name', ''))
        review_transactions.sort(key=lambda x: x[0].get('file_name', ''))
        
        # Add OK transactions first with beautiful styling
        row_number = 0
        for transaction, review_item in ok_transactions:
            file_name = transaction.get('file_name', '')
            row_number += 1
            
            # Determine row background (subtle alternating)
            row_bg = self.colors['row_even'] if row_number % 2 == 0 else self.colors['row_odd']
            
            # Basic transaction data with styling
            cells = [
                ws.cell(row=current_row, column=1, value=file_name),
                ws.cell(row=current_row, column=2, value=transaction.get('date', '')),
                ws.cell(row=current_row, column=3, value=transaction.get('amount', 0)),
                ws.cell(row=current_row, column=4, value=transaction.get('category', 'Other')),
                ws.cell(row=current_row, column=5, value=transaction.get('description', '')),
            ]
            
            # Apply beautiful row styling to data cells
            for cell in cells:
                cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
                cell.border = self.styles['thin_border']
                cell.alignment = Alignment(vertical="center")
            
            # Beautiful OK status cell - emerald green elegance
            ok_cell = ws.cell(row=current_row, column=6, value="✓ OK")
            ok_cell.font = Font(bold=True, color=self.colors['ok_text'])
            ok_cell.fill = PatternFill(start_color=self.colors['ok_bg'], end_color=self.colors['ok_bg'], fill_type="solid")
            ok_cell.alignment = Alignment(horizontal="center", vertical="center")
            ok_cell.border = self.styles['thin_border']
            
            # Empty review cells with row styling
            review_cells = [
                ws.cell(row=current_row, column=7, value=""),
                ws.cell(row=current_row, column=8, value="")
            ]
            for cell in review_cells:
                cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
                cell.border = self.styles['thin_border']
            
            current_row += 1
        
        # Add REVIEW transactions at the end with sophisticated styling
        for transaction, review_item in review_transactions:
            file_name = transaction.get('file_name', '')
            row_number += 1
            
            # Determine row background (continuing alternating pattern)
            row_bg = self.colors['row_even'] if row_number % 2 == 0 else self.colors['row_odd']
            
            # Basic transaction data with styling
            cells = [
                ws.cell(row=current_row, column=1, value=file_name),
                ws.cell(row=current_row, column=2, value=transaction.get('date', '')),
                ws.cell(row=current_row, column=3, value=transaction.get('amount', 0)),
                ws.cell(row=current_row, column=4, value=transaction.get('category', 'Other')),
                ws.cell(row=current_row, column=5, value=transaction.get('description', '')),
            ]
            
            # Apply beautiful row styling to data cells
            for cell in cells:
                cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
                cell.border = self.styles['thin_border']
                cell.alignment = Alignment(vertical="center")
            
            # Stunning REVIEW status cell - warm amber sophistication
            review_cell = ws.cell(row=current_row, column=6, value="⚠ REVIEW")
            review_cell.font = Font(bold=True, color=self.colors['review_text'])
            review_cell.fill = PatternFill(start_color=self.colors['review_bg'], end_color=self.colors['review_bg'], fill_type="solid")
            review_cell.alignment = Alignment(horizontal="center", vertical="center")
            review_cell.border = self.styles['thin_border']
            
            # Review information cells with elegant styling
            reason_cell = ws.cell(row=current_row, column=7, value=review_item.reason)
            reason_cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
            reason_cell.border = self.styles['thin_border']
            reason_cell.alignment = Alignment(vertical="center")
            reason_cell.font = Font(italic=True)
            
            snippet_cell = ws.cell(row=current_row, column=8, value=review_item.raw_snippet)
            snippet_cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
            snippet_cell.border = self.styles['thin_border']
            snippet_cell.alignment = Alignment(vertical="center", wrap_text=True)
            snippet_cell.font = Font(size=9, color='666666')
            
            current_row += 1
        
        # Add any review items that don't have corresponding transactions
        for item in review_items:
            # Clean the filename properly - remove hash suffix and use .pdf extension
            json_file_name = Path(item.file_path).stem
            if '_' in json_file_name:
                file_name = '_'.join(json_file_name.split('_')[:-1]) + '.pdf'
            else:
                file_name = json_file_name + '.pdf'
            # Check if this review item already has a transaction
            has_transaction = any(t.get('file_name') == file_name for t in transactions)
            
            if not has_transaction:
                ws.cell(row=current_row, column=1, value=file_name)
                ws.cell(row=current_row, column=2, value=item.suggested_date or '')
                ws.cell(row=current_row, column=3, value=item.suggested_amount or '')
                ws.cell(row=current_row, column=4, value=item.suggested_category or '')
                ws.cell(row=current_row, column=5, value="")
                ws.cell(row=current_row, column=6, value="REVIEW")
                ws.cell(row=current_row, column=7, value=item.reason)
                ws.cell(row=current_row, column=8, value=item.raw_snippet)
                current_row += 1
        
        # Auto-adjust column widths
        column_widths = [25, 12, 12, 20, 40, 12, 40, 60]
        for i, width in enumerate(column_widths, 1):
            if i <= len(headers):
                column_letter = chr(64 + i)  # Convert to column letter (A, B, C...)
                ws.column_dimensions[column_letter].width = width
        
        logger.info(f"Created consolidated sheet with {len(transactions)} transactions and {len(review_items)} review items")
    
    
    def _add_summary_section(self, ws, transactions: List[Dict[str, Any]], start_row: int) -> int:
        """Add summary statistics to the top of the consolidated sheet."""
        if not transactions:
            ws.cell(row=start_row, column=1, value="No transactions to summarize")
            return start_row + 1
        
        # Convert to DataFrame for easier analysis
        df = pd.DataFrame(transactions)
        
        # Determine month/year from transactions for title
        month_year = self._determine_period_from_transactions(transactions)
        title = f"TRANSACTION SUMMARY - {month_year.upper()}"
        
        # Stunning title with sophisticated styling
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = Font(bold=True, size=18, color=self.colors['summary_accent'])
        title_cell.fill = PatternFill(start_color=self.colors['summary_bg'], end_color=self.colors['summary_bg'], fill_type="solid")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        title_cell.border = self.styles['thin_border']
        # Merge title across multiple columns for impact
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
        current_row = start_row + 2
        
        # Beautiful summary stats with professional layout
        # Total Transactions
        label_cell = ws.cell(row=current_row, column=1, value="Total Transactions:")
        label_cell.font = Font(bold=True, color=self.colors['summary_accent'])
        label_cell.fill = PatternFill(start_color=self.colors['summary_bg'], end_color=self.colors['summary_bg'], fill_type="solid")
        label_cell.border = self.styles['thin_border']
        
        value_cell = ws.cell(row=current_row, column=2, value=len(transactions))
        value_cell.font = Font(bold=True, size=12)
        value_cell.border = self.styles['thin_border']
        value_cell.alignment = Alignment(horizontal="center")
        
        if 'amount' in df.columns:
            total_amount = df['amount'].sum()
            # Total Amount
            total_label = ws.cell(row=current_row, column=4, value="Total Amount:")
            total_label.font = Font(bold=True, color=self.colors['summary_accent'])
            total_label.fill = PatternFill(start_color=self.colors['summary_bg'], end_color=self.colors['summary_bg'], fill_type="solid")
            total_label.border = self.styles['thin_border']
            
            total_value = ws.cell(row=current_row, column=5, value=f"¥{total_amount:,}")
            total_value.font = Font(bold=True, size=12, color='D97706')  # Rich orange for money
            total_value.border = self.styles['thin_border']
            total_value.alignment = Alignment(horizontal="center")
            
            # Average Amount
            avg_amount = df['amount'].mean()
            avg_label = ws.cell(row=current_row, column=7, value="Average Amount:")
            avg_label.font = Font(bold=True, color=self.colors['summary_accent'])
            avg_label.fill = PatternFill(start_color=self.colors['summary_bg'], end_color=self.colors['summary_bg'], fill_type="solid")
            avg_label.border = self.styles['thin_border']
            
            avg_value = ws.cell(row=current_row, column=8, value=f"¥{avg_amount:,.0f}")
            avg_value.font = Font(bold=True, size=12, color='DC2626')  # Elegant red for averages
            avg_value.border = self.styles['thin_border']
            avg_value.alignment = Alignment(horizontal="center")
        
        current_row += 2
        
        # Elegant category breakdown section
        if 'category' in df.columns:
            # Section title with sophisticated styling
            section_title = ws.cell(row=current_row, column=1, value="TOP CATEGORIES BREAKDOWN")
            section_title.font = Font(bold=True, size=12, color=self.colors['summary_accent'])
            section_title.fill = PatternFill(start_color=self.colors['summary_bg'], end_color=self.colors['summary_bg'], fill_type="solid")
            section_title.border = self.styles['thin_border']
            section_title.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            current_row += 1
            
            category_summary = df.groupby('category').agg({
                'amount': ['count', 'sum']
            }).round(0)
            
            # Beautiful mini-headers for category table
            headers = ["Category", "Count", "Amount"]
            for col, header in enumerate(headers, 1):
                header_cell = ws.cell(row=current_row, column=col, value=header)
                header_cell.font = Font(bold=True, size=10, color=self.colors['header_text'])
                header_cell.fill = PatternFill(start_color=self.colors['header_bg'], end_color=self.colors['header_bg'], fill_type="solid")
                header_cell.border = self.styles['thin_border']
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
            
            # Data with alternating row colors - limit to top categories
            top_categories = category_summary.nlargest(5, ('amount', 'sum'))
            for idx, (category, data) in enumerate(top_categories.iterrows()):
                row_bg = self.colors['row_even'] if idx % 2 == 0 else self.colors['row_odd']
                
                # Category name
                cat_cell = ws.cell(row=current_row, column=1, value=category)
                cat_cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
                cat_cell.border = self.styles['thin_border']
                cat_cell.font = Font(bold=True, size=10)
                
                # Count
                count_cell = ws.cell(row=current_row, column=2, value=int(data[('amount', 'count')]))
                count_cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
                count_cell.border = self.styles['thin_border']
                count_cell.alignment = Alignment(horizontal="center")
                
                # Amount with rich color
                amount_cell = ws.cell(row=current_row, column=3, value=f"¥{int(data[('amount', 'sum')]):,}")
                amount_cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
                amount_cell.border = self.styles['thin_border']
                amount_cell.alignment = Alignment(horizontal="right")
                amount_cell.font = Font(bold=True, color='059669')  # Rich teal for amounts
                
                current_row += 1
        
        return current_row
    
    def _determine_period_from_transactions(self, transactions: List[Dict[str, Any]]) -> str:
        """Determine the most common month/year from transaction dates."""
        from collections import Counter
        
        if not transactions:
            return "Unknown Period"
        
        # Extract month/year from valid dates
        month_years = []
        for transaction in transactions:
            date_str = transaction.get('date')
            if date_str:
                try:
                    # Parse ISO date format (YYYY-MM-DD)
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    month_year = date_obj.strftime('%B %Y')  # e.g., "June 2025"
                    month_years.append(month_year)
                except ValueError:
                    continue
        
        if not month_years:
            return "Unknown Period"
        
        # Find most common month/year
        month_year_counts = Counter(month_years)
        most_common = month_year_counts.most_common(1)[0]
        
        # If the most common month/year represents >50% of transactions, use it
        if most_common[1] / len(month_years) > 0.5:
            return most_common[0]
        else:
            return "Mixed Months"
    
    @staticmethod
    def validate_transaction_data(transactions: List[Dict[str, Any]]) -> List[str]:
        """
        Validate transaction data before export.
        
        Args:
            transactions: List of transaction dictionaries
            
        Returns:
            List of validation error messages
        """
        errors = []
        required_fields = ['date', 'amount', 'category', 'description']
        
        for i, transaction in enumerate(transactions):
            for field in required_fields:
                if field not in transaction:
                    errors.append(f"Transaction {i+1}: Missing '{field}' field")
                elif transaction[field] is None:
                    errors.append(f"Transaction {i+1}: '{field}' is None")
                elif field == 'amount' and not isinstance(transaction[field], (int, float)):
                    errors.append(f"Transaction {i+1}: Amount must be numeric")
        
        return errors
    
    @staticmethod
    def create_transaction_dict(date: Optional[str], 
                              amount: Optional[int], 
                              category: str, 
                              description: str,
                              file_path: str = "") -> Dict[str, Any]:
        """
        Create a properly formatted transaction dictionary.
        
        Args:
            date: ISO date string (YYYY-MM-DD)
            amount: Amount in JPY
            category: Category name
            description: Description text
            file_path: Source file path (for metadata)
            
        Returns:
            Transaction dictionary
        """
        return {
            'date': date or '',
            'amount': amount or 0,
            'category': category,
            'description': description,
            'file_name': Path(file_path).name if file_path else ''
        }